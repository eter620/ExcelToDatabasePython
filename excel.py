import openpyxl

"""
Creates the basic functions for the excel classes
Function Descriptions:
changeExcelFile: sets current excel file to the one passed in and sets the worksheet to the active worksheet
getExcel: returns the current workbook
getSheetName: returns the names of all the sheets
setActiveSheet: sets the active sheet to sheet passed in
changeCurrentSheetName: changes active sheets name
createNewWorkSheet: creates a new worksheet
"""

class Excel:

    def __init__(self):
        self._wb = 0
        self._ws = 0
    """
    sets the excel file to the one passed in and sets the work sheet to the last active tab
    """
    def changeExcelFile(self,excelName):
        try:
            self._wb = openpyxl.load_workbook(excelName)
            self._ws = self._wb.active
        except:
            print("File Not Found")

    """
    returns the sheet names
    """
    def getSheetName(self):
        try:
            return self._wb.sheetnames
        except:
            print("An error occurred")


    """
    sets the active sheet to the sheet name sent in
    """
    def setActiveSheet(self, sheetName):
        try:
            self._ws = self._wb[sheetName]
        except:
            print("An error occurred")

    """
    returns the current work book 
    """
    def getExcel(self):
        return self._wb

    """
    changes the name of the active sheet
    """
    def changeCurrentSheetName(self, sheetName):
        self._ws.title = sheetName

    """
    creates a new work sheet 
    """
    def createNewWorkSheet(self, wsName):
        self._wb.create_sheet(wsName)
